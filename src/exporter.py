"""导出 Excel：作品全量（转赞评+播放）、粉丝/获赞趋势、作品逐小时、榜单。

用法（server 端调用）：
    from exporter import build_workbook_bytes
    data: bytes = build_workbook_bytes()                # 默认：张真源全量（四表）
    data: bytes = build_workbook_bytes(sec_uid, nick)   # 一次性抓取某账号（概览+作品）
"""
import io
from datetime import datetime, timezone, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from storage import (get_conn, post_list, latest_post_stats,
                     profile_series, board_latest, post_stats_rows,
                     get_user)

HEADER_FILL = PatternFill("solid", fgColor="16A34A")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN = Side(style="thin", color="E7F7EC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUM_FMT = "#,##0"


def _unix_to_str(u):
    if not u:
        return ""
    try:
        return datetime.fromtimestamp(
            int(u), tz=timezone(timedelta(hours=8))
        ).strftime("%Y-%m-%d %H:%M")
    except Exception:
        return str(u)


def _style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def _fill_works(ws, sec_uid=None):
    """作品全量数据：按点赞降序。"""
    ws.title = "作品数据"
    headers = ["序号", "作品ID", "标题", "发布时间", "点赞", "评论",
               "收藏", "分享", "播放", "分享链接", "封面链接"]
    ws.append(headers)
    _style_header(ws, len(headers))

    posts = post_list(sec_uid)
    enriched = []
    for p in posts:
        aweme_id, desc, create_time, cover, share, play, download, last = p
        latest = latest_post_stats(aweme_id)
        digg = latest[1] if latest else 0
        enriched.append((aweme_id, desc, create_time, cover, share,
                         play, download, latest))
    enriched.sort(key=lambda x: (x[7][1] if x[7] else 0), reverse=True)

    for i, (aweme_id, desc, create_time, cover, share, play, download, latest) in enumerate(enriched, 1):
        if latest:
            digg, comment, collect, sh, pl = latest[1], latest[2], latest[3], latest[4], latest[5]
        else:
            digg = comment = collect = sh = pl = 0
        ws.append([i, aweme_id, (desc or "")[:300], _unix_to_str(create_time),
                   digg, comment, collect, sh, pl, share or "", cover or ""])
        r = ws.max_row
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = BORDER
        for c in (5, 6, 7, 8, 9):
            ws.cell(row=r, column=c).number_format = NUM_FMT

    for i, w in enumerate([6, 22, 42, 18, 12, 10, 10, 10, 12, 28, 38], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"


def _fill_profile(ws, sec_uid=None):
    """粉丝 / 获赞 趋势（按账号过滤）。"""
    ws.title = "粉丝获赞趋势"
    headers = ["时间", "粉丝数", "总获赞", "作品数"]
    ws.append(headers)
    _style_header(ws, len(headers))
    for r in profile_series(sec_uid):
        ts, follower, liked, aweme = r
        ws.append([ts, follower, liked, aweme or ""])
        rr = ws.max_row
        for c in range(1, len(headers) + 1):
            ws.cell(row=rr, column=c).border = BORDER
        ws.cell(row=rr, column=2).number_format = NUM_FMT
        ws.cell(row=rr, column=3).number_format = NUM_FMT
    for i, w in enumerate([22, 14, 16, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _fill_post_hourly(ws, sec_uid=None):
    """作品逐小时互动明细（每条作品每个采集点一行）。"""
    ws.title = "作品逐小时"
    headers = ["时间", "作品ID", "标题", "点赞", "评论", "收藏", "分享", "播放"]
    ws.append(headers)
    _style_header(ws, len(headers))

    title_map = {p[0]: (p[1] or "")[:60] for p in post_list(sec_uid)}
    for ts, aid, digg, comment, collect, sh, pl in post_stats_rows(sec_uid):
        ws.append([ts, aid, title_map.get(aid, ""), digg or 0, comment or 0,
                   collect or 0, sh or 0, pl or 0])
        rr = ws.max_row
        for c in range(1, len(headers) + 1):
            ws.cell(row=rr, column=c).border = BORDER
        for c in (4, 5, 6, 7, 8):
            ws.cell(row=rr, column=c).number_format = NUM_FMT
    for i, w in enumerate([22, 22, 42, 12, 10, 10, 10, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"


def _fill_overview(ws, sec_uid, nickname=None):
    """一次性抓取：账号概览（粉丝/获赞数/作品数等）。"""
    ws.title = "账号概览"
    u = get_user(sec_uid) or {}
    nick = nickname or u.get("nickname") or ""
    rows = [
        ("账号昵称", nick),
        ("抖音号 (unique_id)", u.get("unique_id") or "—"),
        ("sec_uid", sec_uid),
        ("主页链接", u.get("profile_url")
         or f"https://www.douyin.com/user/{sec_uid}"),
        ("粉丝数", u.get("follower_count") or 0),
        ("总获赞", u.get("total_favorited") or 0),
        ("作品数（主页显示）", u.get("aweme_count") or 0),
        ("已抓取作品条数", u.get("post_count") or 0),
        ("抓取时间", u.get("fetched_at") or ""),
    ]
    ws.append(["项目", "内容"])
    _style_header(ws, 2)
    for k, v in rows:
        ws.append([k, v])
        r = ws.max_row
        for c in (1, 2):
            ws.cell(row=r, column=c).border = BORDER
        if k in ("粉丝数", "总获赞", "作品数（主页显示）", "已抓取作品条数"):
            ws.cell(row=r, column=2).number_format = NUM_FMT
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 64
    ws.freeze_panes = "A2"


def _fill_boards(ws):
    """最新一次快照的榜单：挑战榜 / 娱乐热榜 / 张真源话题。"""
    ws.title = "榜单"
    headers = ["类型", "排名", "名称", "数据"]
    ws.append(headers)
    _style_header(ws, len(headers))
    label = {"challenge": "挑战榜", "hot_total": "娱乐热榜", "topic": "张真源话题"}
    for bt in ("challenge", "hot_total", "topic"):
        for it in board_latest(bt):
            extra = " · ".join(f"{k}:{v}" for k, v in (it["extra"] or {}).items())
            ws.append([label.get(bt, bt), it["rank"], it["name"], extra])
            rr = ws.max_row
            for c in range(1, len(headers) + 1):
                ws.cell(row=rr, column=c).border = BORDER
    for i, w in enumerate([14, 8, 40, 50], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def build_workbook_bytes(sec_uid=None, nickname=None):
    wb = Workbook()
    if sec_uid:
        # 一次性抓取：账号概览 + 作品数据
        _fill_overview(wb.active, sec_uid, nickname)
        _fill_works(wb.create_sheet("作品数据"), sec_uid)
    else:
        # 默认：张真源全量（四表）
        _fill_works(wb.active)
        _fill_profile(wb.create_sheet("粉丝获赞趋势"))
        _fill_post_hourly(wb.create_sheet("作品逐小时"))
        _fill_boards(wb.create_sheet("榜单"))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
